#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Export Cash Flow Statement to Excel

Usage:
    python export_to_excel.py test_aapl.csv
    python export_to_excel.py test_wmt_corrected.csv
"""

import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


def format_millions(value):
    """Format value in millions/billions USD"""
    if pd.isna(value):
        return "N/A"
    if value == 0:
        return "$0.00M"
    # Convert to millions
    millions = value / 1_000_000
    if abs(millions) >= 1000:
        billions = millions / 1000
        return f"${billions:.2f}B"
    return f"${millions:.2f}M"


def format_usd(value):
    """Format value as USD"""
    if pd.isna(value):
        return "N/A"
    return f"${value:,.0f}"


def format_excel(ws, df):
    """Format Excel worksheet with styling"""
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Title style
    title_font = Font(bold=True, size=14)
    
    # Data style
    data_font = Font(size=10)
    number_alignment = Alignment(horizontal="right", vertical="center")
    text_alignment = Alignment(horizontal="left", vertical="center")
    
    # Border
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set column widths
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    
    # Title row
    ws.merge_cells('A1:C1')
    title_cell = ws['A1']
    title_cell.value = f"{df['symbol'].iloc[0]} - Cash Flow Statement"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Subtitle row
    ws.merge_cells('A2:C2')
    subtitle_cell = ws['A2']
    subtitle_cell.value = f"Extracted from SEC XBRL Data (FY {df['fiscal_year_end_date'].iloc[0]})"
    subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Basic Information section
    row = 4
    ws[f'A{row}'] = "Basic Information"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    row += 1
    
    info_items = [
        ("Symbol", df['symbol'].iloc[0]),
        ("Form Type", df['form_type'].iloc[0]),
        ("Fiscal Year End", df['fiscal_year_end_date'].iloc[0]),
        ("Filing Date", df['filing_date'].iloc[0]),
        ("Accession Number", df['accession_number'].iloc[0]),
        ("Currency", df['currency'].iloc[0]),
    ]
    
    for label, value in info_items:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        ws[f'A{row}'].alignment = text_alignment
        ws[f'B{row}'].alignment = text_alignment
        row += 1
    
    row += 1
    
    # Cash Flow Statement Data section
    ws[f'A{row}'] = "Cash Flow Statement Data"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    row += 1
    
    # Header row
    headers = ["Item", "Amount (Millions USD)", "Raw Value (USD)"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    row += 1
    
    # Data rows
    cash_flow_items = [
        ("Consolidated Net Income", df['net_income'].iloc[0]),
        ("Net Cash from Operating Activities", df['net_cash_operating'].iloc[0]),
        ("Net Cash from Investing Activities", df['net_cash_investing'].iloc[0]),
        ("Net Cash from Financing Activities", df['net_cash_financing'].iloc[0]),
        ("Effect of Exchange Rates on Cash", df['effect_of_exchange_rates_on_cash'].iloc[0]),
        ("Net Change in Cash", df['net_change_in_cash'].iloc[0]),
        ("Cash Beginning of Period", df['cash_beginning_of_period'].iloc[0]),
        ("Cash End of Period", df['cash_end_of_period'].iloc[0]),
    ]
    
    for item_name, value in cash_flow_items:
        ws[f'A{row}'] = item_name
        ws[f'A{row}'].font = data_font
        ws[f'A{row}'].alignment = text_alignment
        ws[f'A{row}'].border = thin_border
        
        ws[f'B{row}'] = format_millions(value)
        ws[f'B{row}'].font = data_font
        ws[f'B{row}'].alignment = number_alignment
        ws[f'B{row}'].border = thin_border
        
        ws[f'C{row}'] = format_usd(value)
        ws[f'C{row}'].font = data_font
        ws[f'C{row}'].alignment = number_alignment
        ws[f'C{row}'].border = thin_border
        
        row += 1
    
    row += 1
    
    # Key Metrics section
    ws[f'A{row}'] = "Key Metrics"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    row += 1
    
    net_income = df['net_income'].iloc[0]
    net_cash_operating = df['net_cash_operating'].iloc[0]
    net_cash_investing = df['net_cash_investing'].iloc[0]
    
    if pd.notna(net_income) and net_income != 0:
        cf_quality = net_cash_operating / net_income
        ws[f'A{row}'] = "Cash Flow Quality (Operating CF / Consolidated Net Income)"
        ws[f'A{row}'].font = data_font
        ws[f'A{row}'].alignment = text_alignment
        ws[f'B{row}'] = f"{cf_quality:.2f}x"
        ws[f'B{row}'].font = data_font
        ws[f'B{row}'].alignment = number_alignment
        row += 1
    
    if pd.notna(net_cash_operating) and pd.notna(net_cash_investing):
        fcf = net_cash_operating + net_cash_investing
        ws[f'A{row}'] = "Free Cash Flow (Operating CF + Investing CF)"
        ws[f'A{row}'].font = data_font
        ws[f'A{row}'].alignment = text_alignment
        ws[f'B{row}'] = format_millions(fcf)
        ws[f'B{row}'].font = data_font
        ws[f'B{row}'].alignment = number_alignment
    
    # Footer
    row += 2
    ws[f'A{row}'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws[f'A{row}'].font = Font(size=9, italic=True, color="808080")


def export_to_excel(csv_file: str, output_file: str = None):
    """Export CSV to formatted Excel file"""
    df = pd.read_csv(csv_file)
    
    if output_file is None:
        output_file = csv_file.replace('.csv', '.xlsx')
    
    # Create Excel writer
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Raw Data', index=False)
    
    # Load workbook and format
    wb = load_workbook(output_file)
    
    # Create formatted sheet
    ws = wb.create_sheet('Cash Flow Statement', 0)
    format_excel(ws, df)
    
    # Save
    wb.save(output_file)
    print(f"✅ Exported to: {output_file}")
    print(f"   - Sheet 1: Cash Flow Statement (formatted)")
    print(f"   - Sheet 2: Raw Data")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python export_to_excel.py <csv_file> [output_file.xlsx]")
        sys.exit(1)
    
    csv_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    export_to_excel(csv_file, output_file)
