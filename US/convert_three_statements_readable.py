#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Convert US 3-statement workbook to a collaborator-friendly readable workbook.

Usage:
  ./venv/bin/python US/convert_three_statements_readable.py \
    --in US/aapl_3statements.xlsx \
    --out US/check_us_aapl.xlsx
"""

from __future__ import annotations

import argparse
from datetime import datetime
from typing import List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Convert US 3-statement workbook to readable format.")
    p.add_argument("--in", dest="input_path", required=True, help="Input US workbook path")
    p.add_argument("--out", dest="output_path", required=True, help="Output readable workbook path")
    p.add_argument("--include-raw", action="store_true", help="Append raw sheets for internal traceability")
    return p.parse_args()


def is_na(v) -> bool:
    return pd.isna(v) or v is None


def to_float(v) -> Optional[float]:
    if is_na(v):
        return None
    try:
        return float(v)
    except Exception:
        return None


def fmt_short_currency(v, currency: str = "USD") -> str:
    x = to_float(v)
    if x is None:
        return "N/A"
    sign = "-" if x < 0 else ""
    a = abs(x)
    unit = ""
    val = a
    if a >= 1_000_000_000_000:
        val = a / 1_000_000_000_000
        unit = "T"
    elif a >= 1_000_000_000:
        val = a / 1_000_000_000
        unit = "B"
    elif a >= 1_000_000:
        val = a / 1_000_000
        unit = "M"
    symbol = "$" if currency == "USD" else ""
    return f"{symbol}{sign}{val:.2f}{unit}"


def fmt_raw_currency(v, currency: str = "USD") -> str:
    x = to_float(v)
    if x is None:
        return "N/A"
    symbol = "$" if currency == "USD" else ""
    return f"{symbol}{x:,.0f}"


def fmt_number(v) -> str:
    x = to_float(v)
    if x is None:
        return "N/A"
    if abs(x) >= 1000:
        return f"{x:,.0f}"
    return f"{x:,.4f}".rstrip("0").rstrip(".")


def add_sheet_common_style(ws):
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 26


def render_statement_sheet(
    wb: Workbook,
    sheet_name: str,
    title: str,
    subtitle: str,
    meta_items: List[Tuple[str, str]],
    data_items: List[Tuple[str, str, str]],
    row: pd.Series,
    metrics: List[Tuple[str, Optional[str]]],
    currency: str = "USD",
    section_basic_label: str = "Basic Information",
    section_data_label: str = "Statement Data",
    section_metrics_label: str = "Key Metrics",
):
    ws = wb.create_sheet(sheet_name)
    add_sheet_common_style(ws)

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    ws.merge_cells("A1:C1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:C2")
    ws["A2"] = subtitle
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    r = 4
    ws[f"A{r}"] = section_basic_label
    ws[f"A{r}"].font = Font(bold=True, size=11)
    r += 1

    for label, value in meta_items:
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(bold=True)
        ws[f"B{r}"] = value
        ws[f"A{r}"].alignment = Alignment(horizontal="left", vertical="center")
        ws[f"B{r}"].alignment = Alignment(horizontal="left", vertical="center")
        r += 1

    r += 1
    ws[f"A{r}"] = section_data_label
    ws[f"A{r}"].font = Font(bold=True, size=11)
    r += 1

    headers = ["Item", f"Amount (Short {currency})", f"Raw Value ({currency})"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=c)
        cell.value = h
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    r += 1

    for label, key, kind in data_items:
        v = row.get(key)
        ws[f"A{r}"] = label
        ws[f"A{r}"].border = thin_border
        ws[f"A{r}"].alignment = Alignment(horizontal="left", vertical="center")

        if kind == "currency":
            short = fmt_short_currency(v, currency=currency)
            raw = fmt_raw_currency(v, currency=currency)
        else:
            short = fmt_number(v)
            raw = fmt_number(v)

        ws[f"B{r}"] = short
        ws[f"C{r}"] = raw
        ws[f"B{r}"].alignment = Alignment(horizontal="right", vertical="center")
        ws[f"C{r}"].alignment = Alignment(horizontal="right", vertical="center")
        ws[f"B{r}"].border = thin_border
        ws[f"C{r}"].border = thin_border
        r += 1

    r += 1
    ws[f"A{r}"] = section_metrics_label
    ws[f"A{r}"].font = Font(bold=True, size=11)
    r += 1
    for label, value in metrics:
        ws[f"A{r}"] = label
        ws[f"B{r}"] = value if value else "N/A"
        ws[f"B{r}"].alignment = Alignment(horizontal="right", vertical="center")
        r += 1

    r += 2
    ws[f"A{r}"] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws[f"A{r}"].font = Font(size=9, italic=True, color="808080")


def to_ratio_str(num: Optional[float], den: Optional[float]) -> Optional[str]:
    if num is None or den is None or den == 0:
        return None
    return f"{(num / den):.2f}x"


def to_pct_str(num: Optional[float], den: Optional[float]) -> Optional[str]:
    if num is None or den is None or den == 0:
        return None
    return f"{(num / den) * 100:.2f}%"


def convert_workbook(input_path: str, output_path: str, include_raw: bool = False) -> None:
    is_df = pd.read_excel(input_path, sheet_name="US_FIN_IS")
    bs_df = pd.read_excel(input_path, sheet_name="US_FIN_BS")
    cf_df = pd.read_excel(input_path, sheet_name="US_FIN_CF")
    is_row = is_df.iloc[0]
    bs_row = bs_df.iloc[0]
    cf_row = cf_df.iloc[0]

    symbol = str(is_row.get("symbol", "US"))
    fy_end = str(is_row.get("fiscal_year_end_date", ""))
    filing = str(is_row.get("filing_date", ""))
    company_id = str(is_row.get("company_id", ""))
    form_type = str(is_row.get("form_type", ""))
    source_id = str(is_row.get("source_filing_id", cf_row.get("accession_number", "")))
    accounting_std = str(is_row.get("accounting_standard", "US_GAAP"))
    currency = str(is_row.get("currency", "USD"))

    wb = Workbook()
    wb.remove(wb.active)

    meta = [
        ("Ticker", symbol),
        ("Company ID (CIK)", company_id),
        ("Form Type", form_type),
        ("Fiscal Year End", fy_end),
        ("Filing Date", filing),
        ("Source Filing ID", source_id),
        ("Currency", currency),
        ("Accounting Standard", accounting_std),
    ]

    is_data_items = [
        ("Total Revenue", "total_revenue", "currency"),
        ("Cost of Revenue", "cost_of_revenue", "currency"),
        ("Gross Profit", "gross_profit", "currency"),
        ("Operating Expenses", "operating_expenses", "currency"),
        ("Operating Income", "operating_income", "currency"),
        ("Non-operating Income/Expense (Net)", "non_operating_income_expense_net", "currency"),
        ("Other Income", "other_income", "currency"),
        ("Income Before Taxes", "income_before_income_taxes", "currency"),
        ("Income Tax Provision", "provision_for_income_taxes", "currency"),
        ("Net Income", "net_income", "currency"),
        ("EPS Basic", "net_income_per_share_basic", "number"),
        ("EPS Diluted", "net_income_per_share_diluted", "number"),
        ("Shares Outstanding Basic", "shares_outstanding_basic", "number"),
        ("Shares Outstanding Diluted", "shares_outstanding_diluted", "number"),
    ]
    is_metrics = [
        (
            "Operating Margin (Operating Income / Revenue)",
            to_pct_str(to_float(is_row.get("operating_income")), to_float(is_row.get("total_revenue"))),
        ),
        (
            "Net Margin (Net Income / Revenue)",
            to_pct_str(to_float(is_row.get("net_income")), to_float(is_row.get("total_revenue"))),
        ),
    ]
    render_statement_sheet(
        wb=wb,
        sheet_name="Income Statement",
        title=f"{symbol} - Income Statement",
        subtitle=f"SEC XBRL Extracted Data (FYE: {fy_end})",
        meta_items=meta,
        data_items=is_data_items,
        row=is_row,
        metrics=is_metrics,
        currency=currency,
        section_basic_label="Basic Information",
        section_data_label="Income Statement Data",
        section_metrics_label="Key Metrics",
    )

    bs_data_items = [
        ("Total Assets", "total_assets", "currency"),
        ("Cash and Cash Equivalents", "cash_and_cash_equivalents", "currency"),
        ("Marketable Securities (Current)", "marketable_securities_current", "currency"),
        ("Marketable Securities (Non-current)", "marketable_securities_noncurrent", "currency"),
        ("Inventories", "inventories", "currency"),
        ("Accounts Receivable", "accounts_receivable", "currency"),
        ("Other Current Assets", "other_current_assets", "currency"),
        ("Total Current Assets", "total_current_assets", "currency"),
        ("PP&E Net", "property_plant_and_equipment_net", "currency"),
        ("Goodwill", "goodwill", "currency"),
        ("Intangible Assets", "intangible_assets", "currency"),
        ("Other Non-current Assets", "other_noncurrent_assets", "currency"),
        ("Accounts Payable", "accounts_payable", "currency"),
        ("Term Debt (Due Within 1 Year)", "short_term_debt", "currency"),
        ("Other Current Liabilities", "other_current_liabilities", "currency"),
        ("Total Current Liabilities", "total_current_liabilities", "currency"),
        ("Term Debt (Due After 1 Year)", "long_term_debt", "currency"),
        ("Other Non-current Liabilities", "other_noncurrent_liabilities", "currency"),
        ("Total Liabilities", "total_liabilities", "currency"),
        ("Total Shareholders' Equity", "total_shareholders_equity", "currency"),
        ("AOCI", "accumulated_other_comprehensive_income_or_loss", "currency"),
        ("Common Stock", "common_stock", "currency"),
        ("Additional Paid-in Capital", "additional_paid_in_capital", "currency"),
        ("Retained Earnings", "retained_earnings", "currency"),
        ("Non-controlling Interests", "noncontrolling_interests", "currency"),
        ("Total Liabilities + Equity", "total_liabilities_and_shareholders_equity", "currency"),
    ]
    bs_metrics = [
        (
            "Current Ratio (Current Assets / Current Liabilities)",
            to_ratio_str(to_float(bs_row.get("total_current_assets")), to_float(bs_row.get("total_current_liabilities"))),
        ),
        (
            "Debt-to-Equity (Total Liabilities / Equity)",
            to_ratio_str(to_float(bs_row.get("total_liabilities")), to_float(bs_row.get("total_shareholders_equity"))),
        ),
    ]
    render_statement_sheet(
        wb=wb,
        sheet_name="Balance Sheet",
        title=f"{symbol} - Balance Sheet",
        subtitle=f"SEC XBRL Extracted Data (FYE: {fy_end})",
        meta_items=meta,
        data_items=bs_data_items,
        row=bs_row,
        metrics=bs_metrics,
        currency=currency,
        section_basic_label="Basic Information",
        section_data_label="Balance Sheet Data",
        section_metrics_label="Key Metrics",
    )

    cf_data_items = [
        ("Net Income", "net_income", "currency"),
        ("Net Cash from Operating", "net_cash_operating", "currency"),
        ("Net Cash from Investing", "net_cash_investing", "currency"),
        ("Net Cash from Financing", "net_cash_financing", "currency"),
        ("FX Impact on Cash", "effect_of_exchange_rates_on_cash", "currency"),
        ("Net Change in Cash", "net_change_in_cash", "currency"),
        ("Cash at Beginning", "cash_beginning_of_period", "currency"),
        ("Cash at End", "cash_end_of_period", "currency"),
    ]
    fcf = None
    if to_float(cf_row.get("net_cash_operating")) is not None and to_float(cf_row.get("net_cash_investing")) is not None:
        fcf = (to_float(cf_row.get("net_cash_operating")) or 0.0) + (to_float(cf_row.get("net_cash_investing")) or 0.0)
    cf_metrics = [
        (
            "Cash Conversion (CFO / Net Income)",
            to_ratio_str(to_float(cf_row.get("net_cash_operating")), to_float(cf_row.get("net_income"))),
        ),
        ("Free Cash Flow (CFO + CFI)", fmt_short_currency(fcf, currency=currency) if fcf is not None else None),
    ]
    render_statement_sheet(
        wb=wb,
        sheet_name="Cash Flow Statement",
        title=f"{symbol} - Cash Flow Statement",
        subtitle=f"SEC XBRL Extracted Data (FYE: {fy_end})",
        meta_items=meta,
        data_items=cf_data_items,
        row=cf_row,
        metrics=cf_metrics,
        currency=currency,
        section_basic_label="Basic Information",
        section_data_label="Cash Flow Data",
        section_metrics_label="Key Metrics",
    )

    wb.save(output_path)

    if include_raw:
        with pd.ExcelWriter(output_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            is_df.to_excel(writer, index=False, sheet_name="Raw US_FIN_IS")
            bs_df.to_excel(writer, index=False, sheet_name="Raw US_FIN_BS")
            cf_df.to_excel(writer, index=False, sheet_name="Raw US_FIN_CF")


def main() -> None:
    args = parse_args()
    convert_workbook(args.input_path, args.output_path, include_raw=args.include_raw)
    print(f"[OK] Readable workbook generated: {args.output_path}")


if __name__ == "__main__":
    main()
