#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Convert CN 3-statement workbook to a collaborator-friendly readable workbook.

Usage:
  ./venv/bin/python CN/convert_three_statements_readable.py \
    --in CN/300750_3statements_from_web.xlsx \
    --out CN/300750_3statements_for_collab_zh_external.xlsx
"""

from __future__ import annotations

import argparse
from datetime import datetime
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Convert CN 3-statement workbook to readable format.")
    p.add_argument("--in", dest="input_path", required=True, help="Input CN workbook path")
    p.add_argument("--out", dest="output_path", required=True, help="Output readable workbook path")
    p.add_argument("--include-raw", action="store_true", help="Append raw canonical sheets for internal traceability")
    return p.parse_args()


def is_na(v) -> bool:
    return pd.isna(v) or v is None


def to_float(v) -> Optional[float]:
    if is_na(v):
        return None
    try:
        return float(v)
    except Exception:
        return None


def fmt_short_currency(v, currency: str = "CNY") -> str:
    x = to_float(v)
    if x is None:
        return "N/A"
    sign = "-" if x < 0 else ""
    a = abs(x)
    unit = ""
    val = a
    if a >= 1_000_000_000_000:
        val = a / 1_000_000_000_000
        unit = "T"
    elif a >= 1_000_000_000:
        val = a / 1_000_000_000
        unit = "B"
    elif a >= 1_000_000:
        val = a / 1_000_000
        unit = "M"
    symbol = "¥" if currency in {"CNY", "RMB"} else ""
    return f"{symbol}{sign}{val:.2f}{unit}"


def fmt_raw_value(v) -> str:
    x = to_float(v)
    if x is None:
        return "N/A"
    if x.is_integer():
        return f"{int(x):,}"
    return f"{x:,.2f}".rstrip("0").rstrip(".")


def fmt_raw_currency(v, currency: str = "CNY") -> str:
    x = to_float(v)
    if x is None:
        return "N/A"
    symbol = "¥" if currency in {"CNY", "RMB"} else ""
    return f"{symbol}{x:,.0f}"


def fmt_number(v) -> str:
    x = to_float(v)
    if x is None:
        return "N/A"
    if abs(x) >= 1000:
        return f"{x:,.0f}"
    return f"{x:,.4f}".rstrip("0").rstrip(".")


def add_sheet_common_style(ws):
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 26


def render_statement_sheet(
    wb: Workbook,
    sheet_name: str,
    title: str,
    subtitle: str,
    meta_items: List[Tuple[str, str]],
    data_items: List[Tuple[str, str, str]],
    row: pd.Series,
    metrics: List[Tuple[str, Optional[str]]],
    currency: str = "CNY",
    reported_unit: str = "",
    section_basic_label: str = "基本信息",
    section_data_label: str = "报表数据",
    section_metrics_label: str = "关键指标",
):
    ws = wb.create_sheet(sheet_name)
    add_sheet_common_style(ws)

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    ws.merge_cells("A1:C1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:C2")
    ws["A2"] = subtitle
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    r = 4
    ws[f"A{r}"] = section_basic_label
    ws[f"A{r}"].font = Font(bold=True, size=11)
    r += 1

    for label, value in meta_items:
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(bold=True)
        ws[f"B{r}"] = value
        ws[f"A{r}"].alignment = Alignment(horizontal="left", vertical="center")
        ws[f"B{r}"].alignment = Alignment(horizontal="left", vertical="center")
        r += 1

    r += 1
    ws[f"A{r}"] = section_data_label
    ws[f"A{r}"].font = Font(bold=True, size=11)
    r += 1

    unit_text = reported_unit.strip() if reported_unit else currency
    headers = ["项目", "数值", "单位"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=c)
        cell.value = h
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    r += 1

    for label, key, kind in data_items:
        v = row.get(key)
        ws[f"A{r}"] = label
        ws[f"A{r}"].border = thin_border
        ws[f"A{r}"].alignment = Alignment(horizontal="left", vertical="center")

        status_key = f"{key}_status"
        status = str(row.get(status_key, "") or "").strip().upper()
        if status == "NOT_APPLICABLE":
            short = "不适用"
            raw = ""
        elif kind == "currency":
            short = fmt_raw_value(v)
            raw = unit_text
        elif kind == "eps":
            short = fmt_number(v)
            raw = "元/股"
        else:
            short = fmt_number(v)
            raw = ""

        ws[f"B{r}"] = short
        ws[f"C{r}"] = raw
        ws[f"B{r}"].alignment = Alignment(horizontal="right", vertical="center")
        ws[f"C{r}"].alignment = Alignment(horizontal="right", vertical="center")
        ws[f"B{r}"].border = thin_border
        ws[f"C{r}"].border = thin_border
        r += 1

    r += 1
    ws[f"A{r}"] = section_metrics_label
    ws[f"A{r}"].font = Font(bold=True, size=11)
    r += 1
    for label, value in metrics:
        ws[f"A{r}"] = label
        ws[f"B{r}"] = value if value else "N/A"
        ws[f"B{r}"].alignment = Alignment(horizontal="right", vertical="center")
        r += 1

    r += 2
    ws[f"A{r}"] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws[f"A{r}"].font = Font(size=9, italic=True, color="808080")


def to_ratio_str(num: Optional[float], den: Optional[float]) -> Optional[str]:
    if num is None or den is None or den == 0:
        return None
    return f"{(num / den):.2f}x"


def to_pct_str(num: Optional[float], den: Optional[float]) -> Optional[str]:
    if num is None or den is None or den == 0:
        return None
    return f"{(num / den) * 100:.2f}%"


def normalize_cn_symbol(v) -> str:
    """
    Keep CN stock code as 6-digit string (preserve leading zeros).
    """
    if v is None or pd.isna(v):
        return "CN"
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    # Remove non-digits and left-pad to 6 for A-share code style.
    d = "".join(ch for ch in s if ch.isdigit())
    if d:
        return d.zfill(6)
    return s


def load_raw_value_map(input_path: str, sheet_name: str) -> Dict[str, float]:
    try:
        df = pd.read_excel(input_path, sheet_name=sheet_name)
    except Exception:
        return {}
    if df.empty or "item_code" not in df.columns or "value" not in df.columns:
        return {}
    out: Dict[str, float] = {}
    for _, r in df.iterrows():
        code = str(r.get("item_code", "")).strip()
        if not code:
            continue
        v = to_float(r.get("value"))
        if v is not None:
            out[code] = v
    return out


def fill_cn_is_display_row(is_row: pd.Series, raw_is_map: Dict[str, float]) -> pd.Series:
    row = is_row.copy()

    def pick(*codes: str) -> Optional[float]:
        for c in codes:
            v = raw_is_map.get(c)
            if v is not None:
                return v
        return None

    # Prefer raw statement layer for presentation semantics.
    mapping = {
        "total_revenue": ("BIZTOTINCO", "BIZINCO"),
        "cost_of_revenue": ("BIZTOTCOST", "BIZCOST"),
        "taxes_and_surcharges": ("BIZTAX",),
        "selling_expense": ("SALESEXPE",),
        "admin_expense": ("MANAEXPE",),
        "rnd_expense": ("DEVEEXPE",),
        "finance_expense": ("FINEXPE", "FININCO"),
        "net_income": ("NETPROFIT", "PARENETP"),
        "net_income_per_share_basic": ("BASICEPS",),
        "net_income_per_share_diluted": ("DILUTEDEPS",),
    }

    for key, codes in mapping.items():
        v = pick(*codes)
        if v is not None:
            row[key] = v
    return row


def convert_workbook(input_path: str, output_path: str, include_raw: bool = False) -> None:
    is_df = pd.read_excel(input_path, sheet_name="CN_FIN_IS")
    bs_df = pd.read_excel(input_path, sheet_name="CN_FIN_BS")
    cf_df = pd.read_excel(input_path, sheet_name="CN_FIN_CF")

    is_row = is_df.iloc[0]
    bs_row = bs_df.iloc[0]
    cf_row = cf_df.iloc[0]
    raw_is_map = load_raw_value_map(input_path, "RAW_CN_FIN_IS_GEN")
    is_display_row = fill_cn_is_display_row(is_row, raw_is_map)

    symbol = normalize_cn_symbol(is_row.get("symbol", "CN"))
    fy_end = str(is_row.get("fiscal_year_end_date", ""))
    filing = str(is_row.get("filing_date", ""))
    company_id = str(is_row.get("company_id", ""))
    form_type = str(is_row.get("form_type", ""))
    source_id = str(is_row.get("source_filing_id", ""))
    accounting_std = str(is_row.get("accounting_standard", ""))
    currency = str(is_row.get("currency", "CNY"))
    reported_unit = str(is_row.get("reported_unit", "")).strip() or "人民币元"
    finance_label = str(is_row.get("finance_result_source_label", "")).strip()
    if finance_label not in {"财务费用", "财务收入"}:
        finance_label = "财务费用"

    wb = Workbook()
    wb.remove(wb.active)

    meta = [
        ("股票代码", symbol),
        ("公司ID（orgId）", company_id),
        ("报告类型", form_type),
        ("财年截止日", fy_end),
        ("披露日", filing),
        ("源文件ID", source_id),
        ("币种", currency),
        ("报表单位", reported_unit),
        ("会计准则", accounting_std),
    ]

    is_data_items = [
        ("营业总收入", "total_revenue", "currency"),
        ("营业总成本", "cost_of_revenue", "currency"),
        ("税金及附加", "taxes_and_surcharges", "currency"),
        ("销售费用", "selling_expense", "currency"),
        ("管理费用", "admin_expense", "currency"),
        ("研发费用", "rnd_expense", "currency"),
        (finance_label, "finance_expense", "currency"),
        ("营业利润", "operating_income", "currency"),
        ("营业外收支净额", "non_operating_income_expense_net", "currency"),
        ("其他收益", "other_income", "currency"),
        ("税前利润", "income_before_income_taxes", "currency"),
        ("所得税费用", "provision_for_income_taxes", "currency"),
        ("净利润", "net_income", "currency"),
        ("基本每股收益", "net_income_per_share_basic", "eps"),
        ("稀释每股收益", "net_income_per_share_diluted", "eps"),
    ]

    is_metrics = [
        ("营业利润率（营业利润 / 营业总收入）", to_pct_str(to_float(is_display_row.get("operating_income")), to_float(is_display_row.get("total_revenue")))),
        ("净利率（净利润 / 营业总收入）", to_pct_str(to_float(is_display_row.get("net_income")), to_float(is_display_row.get("total_revenue")))),
    ]

    render_statement_sheet(
        wb=wb,
        sheet_name="利润表",
        title=f"{symbol} - 利润表",
        subtitle=f"CNINFO 年报抽取数据（财年截止: {fy_end}）",
        meta_items=meta,
        data_items=is_data_items,
        row=is_display_row,
        metrics=is_metrics,
        currency=currency,
        reported_unit=reported_unit,
        section_basic_label="基本信息",
        section_data_label="利润表数据",
        section_metrics_label="关键指标",
    )

    bs_data_items = [
        ("总资产", "total_assets", "currency"),
        ("货币资金", "cash_and_cash_equivalents", "currency"),
        ("短期投资", "short_term_investments", "currency"),
        ("货币资金+短期投资", "total_cash_and_short_term_investments", "currency"),
        ("存货", "inventories", "currency"),
        ("应收款项", "accounts_receivable", "currency"),
        ("其他流动资产", "other_current_assets", "currency"),
        ("流动资产合计", "total_current_assets", "currency"),
        ("固定资产净额", "property_plant_and_equipment_net", "currency"),
        ("商誉", "goodwill", "currency"),
        ("无形资产", "intangible_assets", "currency"),
        ("其他非流动资产", "other_noncurrent_assets", "currency"),
        ("应付账款", "accounts_payable", "currency"),
        ("短期债务", "short_term_debt", "currency"),
        ("其他流动负债", "other_current_liabilities", "currency"),
        ("流动负债合计", "total_current_liabilities", "currency"),
        ("长期债务", "long_term_debt", "currency"),
        ("其他非流动负债", "other_noncurrent_liabilities", "currency"),
        ("总负债", "total_liabilities", "currency"),
        ("归母股东权益", "total_shareholders_equity", "currency"),
        ("其他综合收益累计额", "accumulated_other_comprehensive_income_or_loss", "currency"),
        ("股本", "common_stock", "currency"),
        ("资本公积", "additional_paid_in_capital", "currency"),
        ("留存收益", "retained_earnings", "currency"),
        ("少数股东权益", "noncontrolling_interests", "currency"),
        ("负债和股东权益合计", "total_liabilities_and_shareholders_equity", "currency"),
    ]

    bs_metrics = [
        ("流动比率（流动资产 / 流动负债）", to_ratio_str(to_float(bs_row.get("total_current_assets")), to_float(bs_row.get("total_current_liabilities")))),
        ("资产负债率（总负债 / 总资产）", to_pct_str(to_float(bs_row.get("total_liabilities")), to_float(bs_row.get("total_assets")))),
    ]

    render_statement_sheet(
        wb=wb,
        sheet_name="资产负债表",
        title=f"{symbol} - 资产负债表",
        subtitle=f"CNINFO 年报抽取数据（财年截止: {fy_end}）",
        meta_items=meta,
        data_items=bs_data_items,
        row=bs_row,
        metrics=bs_metrics,
        currency=currency,
        reported_unit=reported_unit,
        section_basic_label="基本信息",
        section_data_label="资产负债表数据",
        section_metrics_label="关键指标",
    )

    cf_data_items = [
        ("净利润", "net_income", "currency"),
        ("经营活动现金流净额", "net_cash_operating", "currency"),
        ("投资活动现金流净额", "net_cash_investing", "currency"),
        ("筹资活动现金流净额", "net_cash_financing", "currency"),
        ("汇率变动影响", "effect_of_exchange_rates_on_cash", "currency"),
        ("现金净增加额", "net_change_in_cash", "currency"),
        ("期初现金余额", "cash_beginning_of_period", "currency"),
        ("期末现金余额", "cash_end_of_period", "currency"),
    ]

    fcf = None
    if to_float(cf_row.get("net_cash_operating")) is not None and to_float(cf_row.get("net_cash_investing")) is not None:
        fcf = (to_float(cf_row.get("net_cash_operating")) or 0.0) + (to_float(cf_row.get("net_cash_investing")) or 0.0)

    cf_metrics = [
        ("现金创造能力（经营CF / 净利润）", to_ratio_str(to_float(cf_row.get("net_cash_operating")), to_float(cf_row.get("net_income")))),
        ("自由现金流（经营CF + 投资CF）", fmt_raw_value(fcf) if fcf is not None else None),
    ]

    render_statement_sheet(
        wb=wb,
        sheet_name="现金流量表",
        title=f"{symbol} - 现金流量表",
        subtitle=f"CNINFO 年报抽取数据（财年截止: {fy_end}）",
        meta_items=meta,
        data_items=cf_data_items,
        row=cf_row,
        metrics=cf_metrics,
        currency=currency,
        reported_unit=reported_unit,
        section_basic_label="基本信息",
        section_data_label="现金流量表数据",
        section_metrics_label="关键指标",
    )

    wb.save(output_path)

    if include_raw:
        with pd.ExcelWriter(output_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            is_df.to_excel(writer, index=False, sheet_name="Raw CN_FIN_IS")
            bs_df.to_excel(writer, index=False, sheet_name="Raw CN_FIN_BS")
            cf_df.to_excel(writer, index=False, sheet_name="Raw CN_FIN_CF")


def main():
    args = parse_args()
    convert_workbook(args.input_path, args.output_path, include_raw=args.include_raw)
    print(f"[OK] Readable workbook generated: {args.output_path}")


if __name__ == "__main__":
    main()
