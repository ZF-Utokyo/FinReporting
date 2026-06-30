#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Convert JP 3-statement raw workbook to a collaborator-friendly readable workbook.

Usage:
  ./venv/bin/python JP/convert_three_statements_readable.py \
    --in JP/toyota_3statements_from_web.xlsx \
    --out JP/toyota_3statements_readable.xlsx
"""

from __future__ import annotations

import argparse
from datetime import datetime
from typing import List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Convert JP 3-statement workbook to readable format.")
    p.add_argument("--in", dest="input_path", required=True, help="Input JP workbook path")
    p.add_argument("--out", dest="output_path", required=True, help="Output readable workbook path")
    p.add_argument("--include-raw", action="store_true", help="Append raw sheets for internal traceability")
    return p.parse_args()


def is_na(v) -> bool:
    return pd.isna(v) or v is None


def to_float(v) -> Optional[float]:
    if is_na(v):
        return None
    try:
        return float(v)
    except Exception:
        return None


def fmt_short_currency(v, currency: str = "JPY") -> str:
    x = to_float(v)
    if x is None:
        return "N/A"
    sign = "-" if x < 0 else ""
    a = abs(x)
    unit = ""
    val = a
    if a >= 1_000_000_000_000:
        val = a / 1_000_000_000_000
        unit = "T"
    elif a >= 1_000_000_000:
        val = a / 1_000_000_000
        unit = "B"
    elif a >= 1_000_000:
        val = a / 1_000_000
        unit = "M"
    symbol = "¥" if currency == "JPY" else ""
    return f"{symbol}{sign}{val:.2f}{unit}"


def fmt_raw_currency(v, currency: str = "JPY") -> str:
    x = to_float(v)
    if x is None:
        return "N/A"
    symbol = "¥" if currency == "JPY" else ""
    return f"{symbol}{x:,.0f}"


def fmt_number(v) -> str:
    x = to_float(v)
    if x is None:
        return "N/A"
    if abs(x) >= 1000:
        return f"{x:,.0f}"
    return f"{x:,.4f}".rstrip("0").rstrip(".")


def add_sheet_common_style(ws):
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 26


def render_statement_sheet(
    wb: Workbook,
    sheet_name: str,
    title: str,
    subtitle: str,
    meta_items: List[Tuple[str, str]],
    data_items: List[Tuple[str, str, str]],
    row: pd.Series,
    metrics: List[Tuple[str, Optional[str]]],
    currency: str = "JPY",
    section_basic_label: str = "基本情報",
    section_data_label: str = "データ",
    section_metrics_label: str = "主要指標",
):
    ws = wb.create_sheet(sheet_name)
    add_sheet_common_style(ws)

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    ws.merge_cells("A1:C1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:C2")
    ws["A2"] = subtitle
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    r = 4
    ws[f"A{r}"] = section_basic_label
    ws[f"A{r}"].font = Font(bold=True, size=11)
    r += 1

    for label, value in meta_items:
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(bold=True)
        ws[f"B{r}"] = value
        ws[f"A{r}"].alignment = Alignment(horizontal="left", vertical="center")
        ws[f"B{r}"].alignment = Alignment(horizontal="left", vertical="center")
        r += 1

    r += 1
    ws[f"A{r}"] = section_data_label
    ws[f"A{r}"].font = Font(bold=True, size=11)
    r += 1

    headers = ["項目", f"金額（簡易表示 {currency}）", f"原数値（{currency}）"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=c)
        cell.value = h
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    r += 1

    for label, key, kind in data_items:
        v = row.get(key)
        ws[f"A{r}"] = label
        ws[f"A{r}"].border = thin_border
        ws[f"A{r}"].alignment = Alignment(horizontal="left", vertical="center")

        if kind == "currency":
            short = fmt_short_currency(v, currency=currency)
            raw = fmt_raw_currency(v, currency=currency)
        else:
            short = fmt_number(v)
            raw = fmt_number(v)

        ws[f"B{r}"] = short
        ws[f"C{r}"] = raw
        ws[f"B{r}"].alignment = Alignment(horizontal="right", vertical="center")
        ws[f"C{r}"].alignment = Alignment(horizontal="right", vertical="center")
        ws[f"B{r}"].border = thin_border
        ws[f"C{r}"].border = thin_border
        r += 1

    r += 1
    ws[f"A{r}"] = section_metrics_label
    ws[f"A{r}"].font = Font(bold=True, size=11)
    r += 1
    for label, value in metrics:
        ws[f"A{r}"] = label
        ws[f"B{r}"] = value if value else "N/A"
        ws[f"B{r}"].alignment = Alignment(horizontal="right", vertical="center")
        r += 1

    r += 2
    ws[f"A{r}"] = f"生成日時: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws[f"A{r}"].font = Font(size=9, italic=True, color="808080")


def to_ratio_str(num: Optional[float], den: Optional[float]) -> Optional[str]:
    if num is None or den is None or den == 0:
        return None
    return f"{(num / den):.2f}x"


def to_pct_str(num: Optional[float], den: Optional[float]) -> Optional[str]:
    if num is None or den is None or den == 0:
        return None
    return f"{(num / den) * 100:.2f}%"


def convert_workbook(input_path: str, output_path: str, include_raw: bool = False) -> None:
    is_df = pd.read_excel(input_path, sheet_name="JP_FIN_IS")
    bs_df = pd.read_excel(input_path, sheet_name="JP_FIN_BS")
    cf_df = pd.read_excel(input_path, sheet_name="JP_FIN_CF")
    is_row = is_df.iloc[0]
    bs_row = bs_df.iloc[0]
    cf_row = cf_df.iloc[0]

    symbol = str(is_row.get("symbol", "JP"))
    fy_end = str(is_row.get("fiscal_year_end_date", ""))
    filing = str(is_row.get("filing_date", ""))
    company_id = str(is_row.get("company_id", ""))
    form_type = str(is_row.get("form_type", ""))
    source_id = str(is_row.get("source_filing_id", ""))
    accounting_std = str(is_row.get("accounting_standard", ""))
    currency = str(is_row.get("currency", "JPY"))

    wb = Workbook()
    wb.remove(wb.active)

    meta = [
        ("銘柄コード", symbol),
        ("会社ID（EDINET）", company_id),
        ("書類種別", form_type),
        ("会計期末日", fy_end),
        ("提出日", filing),
        ("ソース書類ID", source_id),
        ("通貨", currency),
        ("会計基準", accounting_std),
    ]

    is_data_items = [
        ("売上高", "total_revenue", "currency"),
        ("売上原価", "cost_of_revenue", "currency"),
        ("売上総利益", "gross_profit", "currency"),
        ("販売費及び一般管理費", "operating_expenses", "currency"),
        ("営業利益", "operating_income", "currency"),
        ("営業外損益（純額）", "non_operating_income_expense_net", "currency"),
        ("その他収益", "other_income", "currency"),
        ("税引前利益", "income_before_income_taxes", "currency"),
        ("法人税等", "provision_for_income_taxes", "currency"),
        ("当期純利益", "net_income", "currency"),
    ]

    is_metrics = [
        (
            "営業利益率（営業利益 / 売上高）",
            to_pct_str(to_float(is_row.get("operating_income")), to_float(is_row.get("total_revenue"))),
        ),
        (
            "純利益率（当期純利益 / 売上高）",
            to_pct_str(to_float(is_row.get("net_income")), to_float(is_row.get("total_revenue"))),
        ),
    ]

    render_statement_sheet(
        wb=wb,
        sheet_name="損益計算書",
        title=f"{symbol} - 損益計算書",
        subtitle=f"EDINET XBRL 抽出データ（会計期末: {fy_end}）",
        meta_items=meta,
        data_items=is_data_items,
        row=is_row,
        metrics=is_metrics,
        currency=currency,
        section_basic_label="基本情報",
        section_data_label="損益計算書データ",
        section_metrics_label="主要指標",
    )

    bs_data_items = [
        ("資産合計", "total_assets", "currency"),
        ("流動資産合計", "total_current_assets", "currency"),
        ("現金及び預金", "cash_and_cash_equivalents", "currency"),
        ("受取手形・売掛金・契約資産", "accounts_receivable", "currency"),
        ("有価証券", "short_term_investments", "currency"),
        ("棚卸資産", "inventories", "currency"),
        ("その他（流動資産）", "current_assets_other_jp", "currency"),
        ("固定資産合計", "total_noncurrent_assets", "currency"),
        ("有形固定資産合計", "tangible_fixed_assets_total", "currency"),
        ("無形固定資産", "intangible_assets", "currency"),
        ("投資その他の資産合計", "investments_and_other_assets_total", "currency"),
        ("負債純資産合計", "total_liabilities_and_shareholders_equity", "currency"),
        ("負債合計", "total_liabilities", "currency"),
        ("流動負債合計", "total_current_liabilities", "currency"),
        ("支払手形及び買掛金", "accounts_payable", "currency"),
        ("短期借入金", "short_term_debt", "currency"),
        ("その他（流動負債）", "current_liabilities_other_jp", "currency"),
        ("固定負債合計", "total_noncurrent_liabilities", "currency"),
        ("長期借入金", "long_term_debt", "currency"),
        ("その他（固定負債）", "noncurrent_liabilities_other_jp", "currency"),
        ("純資産合計", "total_net_assets", "currency"),
        ("株主資本合計", "total_shareholders_equity", "currency"),
        ("非支配株主持分", "noncontrolling_interests", "currency"),
        ("その他の包括利益累計額合計", "accumulated_other_comprehensive_income_or_loss", "currency"),
        ("新株予約権", "share_subscription_rights", "currency"),
    ]

    bs_metrics = [
        (
            "流動比率（流動資産 / 流動負債）",
            to_ratio_str(to_float(bs_row.get("total_current_assets")), to_float(bs_row.get("total_current_liabilities"))),
        ),
        (
            "D/Eレシオ（負債合計 / 株主資本）",
            to_ratio_str(to_float(bs_row.get("total_liabilities")), to_float(bs_row.get("total_shareholders_equity"))),
        ),
    ]

    render_statement_sheet(
        wb=wb,
        sheet_name="貸借対照表",
        title=f"{symbol} - 貸借対照表",
        subtitle=f"EDINET XBRL 抽出データ（会計期末: {fy_end}）",
        meta_items=meta,
        data_items=bs_data_items,
        row=bs_row,
        metrics=bs_metrics,
        currency=currency,
        section_basic_label="基本情報",
        section_data_label="貸借対照表データ",
        section_metrics_label="主要指標",
    )

    cf_data_items = [
        ("当期純利益", "net_income", "currency"),
        ("営業活動によるキャッシュ・フロー", "net_cash_operating", "currency"),
        ("投資活動によるキャッシュ・フロー", "net_cash_investing", "currency"),
        ("財務活動によるキャッシュ・フロー", "net_cash_financing", "currency"),
        ("為替換算差額", "effect_of_exchange_rates_on_cash", "currency"),
        ("現金及び現金同等物の増減額", "net_change_in_cash", "currency"),
        ("期首現金及び現金同等物残高", "cash_beginning_of_period", "currency"),
        ("期末現金及び現金同等物残高", "cash_end_of_period", "currency"),
    ]

    cf_metrics = [
        (
            "キャッシュ創出力（営業CF / 当期純利益）",
            to_ratio_str(to_float(cf_row.get("net_cash_operating")), to_float(cf_row.get("net_income"))),
        ),
        (
            "フリーキャッシュフロー（営業CF + 投資CF）",
            fmt_short_currency(
                (
                    (to_float(cf_row.get("net_cash_operating")) or 0.0)
                    + (to_float(cf_row.get("net_cash_investing")) or 0.0)
                )
                if to_float(cf_row.get("net_cash_operating")) is not None
                and to_float(cf_row.get("net_cash_investing")) is not None
                else None,
                currency=currency,
            ),
        ),
    ]

    render_statement_sheet(
        wb=wb,
        sheet_name="キャッシュ・フロー計算書",
        title=f"{symbol} - キャッシュ・フロー計算書",
        subtitle=f"EDINET XBRL 抽出データ（会計期末: {fy_end}）",
        meta_items=meta,
        data_items=cf_data_items,
        row=cf_row,
        metrics=cf_metrics,
        currency=currency,
        section_basic_label="基本情報",
        section_data_label="キャッシュ・フローデータ",
        section_metrics_label="主要指標",
    )

    # Write formatted workbook first, then optionally append raw data.
    wb.save(output_path)

    if include_raw:
        with pd.ExcelWriter(output_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            is_df.to_excel(writer, index=False, sheet_name="Raw JP_FIN_IS")
            bs_df.to_excel(writer, index=False, sheet_name="Raw JP_FIN_BS")
            cf_df.to_excel(writer, index=False, sheet_name="Raw JP_FIN_CF")


def main():
    args = parse_args()
    convert_workbook(args.input_path, args.output_path, include_raw=args.include_raw)
    print(f"[OK] Readable workbook generated: {args.output_path}")


if __name__ == "__main__":
    main()
